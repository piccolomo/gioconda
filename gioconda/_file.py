from gioconda._matrix import matrix_class
from gioconda._data import nl
import os, sys
import inspect
import openpyxl
from datetime import datetime

# File Utilities
script_folder = lambda: os.path.abspath(os.path.join(inspect.getfile(sys._getframe(1)), os.pardir))
source_folder = os.path.dirname(os.path.realpath(__file__))
test_data_path = os.path.join(source_folder, 'test_data.csv')

join = os.path.join

def _read_text(path, log = True):
    print("reading text lines in", path) if log else None
    with open(path, 'r', encoding = "utf-8") as file:
        text = file.readlines()
    text = [line.replace('\ufeff', '') for line in text if line != nl]
    print("text lines read!\n") if log else None
    return text

def _split_lines(lines, delimiter = ','):
    return [line.replace("\n", "").split(delimiter) for line in lines]

def _read_csv(path, delimiter = ',', log = True):
    lines = _read_text(path, log = log)
    matrix = _split_lines(lines, delimiter)
    return matrix


def read_csv(file_name, delimiter = ',', header = False, log = True):
    print('loading data') if log else None 
    matrix = _read_csv(file_name, delimiter = delimiter, log = log)
    data = matrix_class()
    data._add_matrix(matrix, header)
    print('data loaded!\n') if log else None
    return data


def _read_xlsx(path, log = True):
    print("reading excel file in", path) if log else None
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active
    matrix = []
    for row in sheet.iter_rows(min_row = 1, max_row = sheet.max_row, min_col = 1, max_col = sheet.max_column):
        line = []
        for cell in row:
            el = cell.value
            el = el.strftime('%d/%m/%Y') if isinstance(el, datetime) else str(el)
            line.append(el)
        matrix.append(line)
    workbook.close()
    return matrix

def read_xlsx(file_name, header = False, log = True):
    print('loading data') if log else None 
    matrix = _read_xlsx(file_name, log = log)
    data = matrix_class()
    data._add_matrix(matrix, header)
    print('data loaded!\n') if log else None
    return data


def read_pickle(path, log = True):
    import pickle
    print("reading pickle", path) if log else None
    with open(path, 'rb') as f:
        data = pickle.load(f)
    print("pickle read!\n") if log else None
    return data
